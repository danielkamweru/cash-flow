"""Render a filtered transaction ledger as XLSX or PDF."""

from __future__ import annotations

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app import models

COLUMNS = ("Date", "Description", "Category", "Type", "Status", "Amount (KES)")

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_BRAND = colors.HexColor("#1F3864")
_STRIPE = colors.HexColor("#F2F5FA")


def _rows(transactions: list[models.Transaction]) -> list[tuple]:
    return [
        (
            t.Date.strftime("%Y-%m-%d") if t.Date else "",
            t.Description or "",
            t.Category or "",
            (t.Type or "").title(),
            (t.Status or "completed").title(),
            round(t.Amount or 0.0, 2),
        )
        for t in transactions
    ]


def _totals(transactions: list[models.Transaction]) -> tuple[float, float]:
    inflow = sum(t.Amount for t in transactions if t.Type == "inflow")
    outflow = sum(t.Amount for t in transactions if t.Type == "outflow")
    return inflow, outflow


def _period(start: date | datetime | None, end: date | datetime | None) -> str:
    fmt = lambda d: d.strftime("%d %b %Y") if d else "—"  # noqa: E731
    return f"{fmt(start)} to {fmt(end)}"


def build_xlsx(
    transactions: list[models.Transaction],
    *,
    entity_name: str,
    start: date | datetime | None,
    end: date | datetime | None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.append([f"{entity_name} — Transaction statement"])
    ws["A1"].font = Font(size=14, bold=True, color="1F3864")
    ws.append([f"Period: {_period(start, end)}"])
    ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(list(COLUMNS))
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in _rows(transactions):
        ws.append(list(row))

    amount_col = get_column_letter(len(COLUMNS))
    for r in range(header_row + 1, ws.max_row + 1):
        ws[f"{amount_col}{r}"].number_format = "#,##0.00"

    inflow, outflow = _totals(transactions)
    ws.append([])
    for label, value in (
        ("Total inflow", inflow),
        ("Total outflow", outflow),
        ("Net", inflow - outflow),
    ):
        ws.append([label, "", "", "", "", round(value, 2)])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=len(COLUMNS)).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=len(COLUMNS)).number_format = "#,##0.00"

    for idx, width in enumerate((14, 46, 18, 12, 13, 16), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = ws[f"A{header_row + 1}"]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf(
    transactions: list[models.Transaction],
    *,
    entity_name: str,
    start: date | datetime | None,
    end: date | datetime | None,
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"{entity_name} transactions",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("wlTitle", parent=styles["Title"], fontSize=17, textColor=_BRAND, alignment=0)
    meta_style = ParagraphStyle("wlMeta", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#5A6377"))
    cell_style = ParagraphStyle("wlCell", parent=styles["Normal"], fontSize=8.5, leading=11)

    story = [
        Paragraph(f"{entity_name} — Transaction statement", title_style),
        Spacer(1, 3 * mm),
        Paragraph(
            f"Period: {_period(start, end)} &nbsp;·&nbsp; {len(transactions)} transactions "
            f"&nbsp;·&nbsp; Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            meta_style,
        ),
        Spacer(1, 5 * mm),
    ]

    data = [list(COLUMNS)]
    for row in _rows(transactions):
        data.append(
            [
                row[0],
                Paragraph(str(row[1]), cell_style),
                row[2],
                row[3],
                row[4],
                f"{row[5]:,.2f}",
            ]
        )

    if len(data) == 1:
        story.append(Paragraph("No transactions in this period.", styles["Normal"]))
    else:
        table = Table(
            data,
            colWidths=[24 * mm, 96 * mm, 38 * mm, 24 * mm, 26 * mm, 32 * mm],
            repeatRows=1,
        )
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), _BRAND),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6DCE8")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        for i in range(1, len(data)):
            if i % 2 == 0:
                style.append(("BACKGROUND", (0, i), (-1, i), _STRIPE))
        table.setStyle(TableStyle(style))
        story.append(table)

        inflow, outflow = _totals(transactions)
        story.append(Spacer(1, 5 * mm))
        summary = Table(
            [
                ["Total inflow", f"{inflow:,.2f}"],
                ["Total outflow", f"{outflow:,.2f}"],
                ["Net", f"{inflow - outflow:,.2f}"],
            ],
            colWidths=[46 * mm, 32 * mm],
            hAlign="RIGHT",
        )
        summary.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("LINEABOVE", (0, 2), (-1, 2), 0.8, _BRAND),
                    ("TEXTCOLOR", (0, 2), (-1, 2), _BRAND),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(summary)

    story.append(Spacer(1, 6 * mm))
    story.append(
        Paragraph(
            "Wealth Loop statement for personal record-keeping. Not a bank statement or tax document.",
            meta_style,
        )
    )

    doc.build(story)
    return buffer.getvalue()
